#!/usr/bin/env python3
"""
Multi-Worksheet Excel Report Generator from Snowflake
Generates Excel workbooks matching template structure with detail and summary tables.
All worksheet structures are hardcoded based on template images.
"""

import argparse
import sys
import os
import yaml
import re
from collections import defaultdict
from dataclasses import dataclass
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import snowflake.connector
except ImportError:
    print("Error: snowflake-connector-python not installed. Run: pip install snowflake-connector-python")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# Configuration Classes
# ============================================================================

@dataclass
class SnowflakeConfig:
    """Snowflake connection configuration"""
    account: str
    user: Optional[str] = None
    password: Optional[str] = None
    warehouse: str = ""
    database: str = ""
    schema: str = ""
    authenticator: str = "externalbrowser"  # 'externalbrowser' for SSO or 'snowflake' for user/pass


@dataclass
class AggregateConfig:
    """Aggregate function configuration"""
    field: str
    function: str  # COUNT, SUM, AVG, etc.
    label: str


@dataclass
class SummaryConfig:
    """Summary table configuration"""
    group_by: str
    aggregates: List[AggregateConfig]
    start_column: str
    columns: List[str]


@dataclass
class FormattingConfig:
    """Formatting configuration"""
    header_row: int = 1
    filters: bool = True
    highlight_columns: List[str] = None
    highlight_cells: List[str] = None
    currency_columns: List[str] = None  # Column names to format as currency (with $ symbol)


@dataclass
class WorksheetConfig:
    """Worksheet configuration"""
    name: str
    query: str
    detail_start_column: str = "A"
    detail_columns: List[str] = None
    spacing_columns: List[str] = None
    summary_config: List[SummaryConfig] = None
    formatting: FormattingConfig = None
    layout_type: str = "table"  # "table" or "form"
    form_layout: Dict[str, Any] = None


# ============================================================================
# Template Type Configuration Functions
# ============================================================================

def create_summary_config_from_template_type(template_type: str, detail_columns_count: int = 7) -> List[SummaryConfig]:
    """
    Create summary configuration based on template type.
    
    Template types:
    - 'direct_dump': No summaries
    - 'direct_dump_state_summary': Detail + Issue State + Resident State summaries
    - 'direct_dump_tat_summary': Detail + TAT summary
    - 'direct_dump_state_tat_summary': Detail + Issue State + Resident State + TAT summaries
    - 'state_summary_only': Only summaries (Issue State + Resident State), no detail
    - 'state_summary_with_company': Only summaries (Issue State + Resident State) with Count and Company columns
    - 'direct_dump_state_payreq_summary': Detail + Issue State + Resident State + Year Pay Req Received summaries
    
    Args:
        template_type: Type of template
        detail_columns_count: Number of detail columns (for spacing calculation)
    
    Returns:
        List of SummaryConfig objects
    """
    summary_configs = []
    
    if template_type == 'direct_dump':
        return None  # No summaries
    
    elif template_type == 'direct_dump_state_summary':
        # Detail + Issue State + Resident State summaries
        # Calculate spacing: detail ends at column (detail_columns_count), gap at (detail_columns_count + 1)
        detail_end_col = chr(ord('A') + detail_columns_count - 1)  # e.g., 'G' for 7 columns
        gap_col = chr(ord('A') + detail_columns_count)  # e.g., 'H'
        issue_start = chr(ord('A') + detail_columns_count + 1)  # e.g., 'I'
        issue_end = chr(ord('A') + detail_columns_count + 2)  # e.g., 'J'
        resident_gap = chr(ord('A') + detail_columns_count + 3)  # e.g., 'K'
        resident_start = chr(ord('A') + detail_columns_count + 4)  # e.g., 'L'
        
        summary_configs = [
            SummaryConfig(
                group_by='Issue_State',
                aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                start_column=issue_start,
                columns=['Issue State', 'Count']
            ),
            SummaryConfig(
                group_by='Resident_State',
                aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                start_column=resident_start,
                columns=['Resident State', 'Count']
            )
        ]
    
    elif template_type == 'direct_dump_tat_summary':
        # Detail + TAT summary
        detail_end_col = chr(ord('A') + detail_columns_count - 1)
        gap_col = chr(ord('A') + detail_columns_count)
        tat_start = chr(ord('A') + detail_columns_count + 1)
        
        summary_configs = [
            SummaryConfig(
                group_by='TAT_Range',
                aggregates=[AggregateConfig(field='TAT_in_Days', function='COUNT', label='TAT COUNTS')],
                start_column=tat_start,
                columns=['', 'TAT COUNTS', '% of Total']
            )
        ]
    
    elif template_type == 'direct_dump_state_tat_summary':
        # Detail + Issue State + Resident State + TAT summaries
        detail_end_col = chr(ord('A') + detail_columns_count - 1)
        gap_col = chr(ord('A') + detail_columns_count)
        tat_start = chr(ord('A') + detail_columns_count + 1)
        tat_end = chr(ord('A') + detail_columns_count + 3)
        issue_gap = chr(ord('A') + detail_columns_count + 4)
        issue_start = chr(ord('A') + detail_columns_count + 5)
        issue_end = chr(ord('A') + detail_columns_count + 6)
        resident_gap = chr(ord('A') + detail_columns_count + 7)
        resident_start = chr(ord('A') + detail_columns_count + 8)
        
        summary_configs = [
            SummaryConfig(
                group_by='TAT_Range',
                aggregates=[AggregateConfig(field='TAT_in_Days', function='COUNT', label='TAT COUNTS')],
                start_column=tat_start,
                columns=['', 'TAT COUNTS', '% of Total']
            ),
            SummaryConfig(
                group_by='Issue_State',
                aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                start_column=issue_start,
                columns=['Issue State', 'CountOfPolicy No']
            ),
            SummaryConfig(
                group_by='Resident_State',
                aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                start_column=resident_start,
                columns=['Resident State', 'CountOfPolicy No']
            )
        ]
    
    elif template_type == 'state_summary_only':
        # Only summaries, no detail
        summary_configs = [
            SummaryConfig(
                group_by='Issue_State',
                aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                start_column='A',
                columns=['Issue State', 'CountOfPolicy No']
            ),
            SummaryConfig(
                group_by='Resident_State',
                aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                start_column='D',  # Gap at C
                columns=['Resident State', 'CountOfPolicy No']
            )
        ]
    
    elif template_type == 'state_summary_with_company':
        # Only summaries with Count and Company columns (for 5-002, 6-001)
        # If detail records have a Count column, use SUM; otherwise COUNT Policy_Num
        # Company always uses COUNT to get distinct company count per state
        summary_configs = [
            SummaryConfig(
                group_by='Issue_State',
                aggregates=[
                    AggregateConfig(field='Count', function='SUM', label='Count'),  # SUM Count column
                    AggregateConfig(field='Company', function='FIRST', label='Company')  # Display company name (first value)
                ],
                start_column='A',  # Issue State: A-B-C
                columns=['Issue State', 'Count', 'Company']
            ),
            SummaryConfig(
                group_by='Resident_State',
                aggregates=[
                    AggregateConfig(field='Count', function='SUM', label='Count'),  # SUM Count column
                    AggregateConfig(field='Company', function='FIRST', label='Company')  # Display company name (first value)
                ],
                start_column='E',  # Gap D, Resident State: E-F-G
                columns=['Resident State', 'Count', 'Company']
            )
        ]
    
    elif template_type == 'direct_dump_state_payreq_summary':
        # Detail + Issue State + Resident State + Year Pay Req Received summaries
        detail_end_col = chr(ord('A') + detail_columns_count - 1)
        gap_col = chr(ord('A') + detail_columns_count)
        issue_start = chr(ord('A') + detail_columns_count + 1)
        issue_end = chr(ord('A') + detail_columns_count + 2)
        resident_gap = chr(ord('A') + detail_columns_count + 3)
        resident_start = chr(ord('A') + detail_columns_count + 4)
        resident_end = chr(ord('A') + detail_columns_count + 5)
        payreq_gap = chr(ord('A') + detail_columns_count + 6)
        payreq_start = chr(ord('A') + detail_columns_count + 7)
        
        summary_configs = [
            SummaryConfig(
                group_by='Issue_State',
                aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                start_column=issue_start,
                columns=['Issue State', 'Count']
            ),
            SummaryConfig(
                group_by='Resident_State',
                aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                start_column=resident_start,
                columns=['Resident State', 'Count']
            ),
            SummaryConfig(
                group_by='Year_Pay_Req_Received',
                aggregates=[AggregateConfig(field='', function='COUNT', label='Counts')],  # Empty field = count records
                start_column=payreq_start,
                columns=['Year Pay Req Received', 'Counts']
            )
        ]
    
    else:
        raise ValueError(f"Unknown template type: {template_type}. Valid types: direct_dump, direct_dump_state_summary, direct_dump_tat_summary, direct_dump_state_tat_summary, state_summary_only, state_summary_with_company, direct_dump_state_payreq_summary")
    
    return summary_configs


def create_worksheet_config_from_template(worksheet_name: str, table_name: str, template_type: str, 
                                         query: Optional[str] = None, detail_columns: Optional[List[str]] = None,
                                         filter_clause: Optional[str] = None,
                                         currency_columns: Optional[List[str]] = None) -> WorksheetConfig:
    """
    Create worksheet configuration from template type.
    
    Args:
        worksheet_name: Name of the worksheet
        table_name: Snowflake table name (used only if query is not provided)
        template_type: Type of template (see create_summary_config_from_template_type)
        query: Optional custom SQL query (if provided, used directly; if not, generates default)
        detail_columns: Optional list of detail column names (if not provided, uses actual columns from query)
        filter_clause: Optional WHERE clause filter (e.g., "Schedule_ID = '1-001' AND Status = 'Active'")
                      Only used if query is not provided
    
    Returns:
        WorksheetConfig object
    """
    # If custom query is provided, use it directly (supports multi-line SQL)
    if query:
        # Clean up query - remove leading/trailing whitespace, normalize newlines
        query = query.strip()
        # Debug: Print the query as read from config (first 300 chars)
        print(f"  DEBUG: Query from config for '{worksheet_name}' (first 300 chars): {query[:300]}")
        # Use the provided query as-is
        pass
    else:
        # Generate default query if not provided
        # Build WHERE clause
        if filter_clause:
            where_clause = f"WHERE {filter_clause}"
        else:
            where_clause = f"WHERE Schedule_ID = '{worksheet_name}'"
        
        if template_type == 'state_summary_only':
            # For summary-only, we still need a query to get data for summaries
            query = f"SELECT Policy_Num, Issue_State, Resident_State FROM {table_name} {where_clause}"
        elif template_type == 'state_summary_with_company':
            # For summary with company, include Count and Company columns
            query = f"SELECT Issue_State, Resident_State, Company, Count FROM {table_name} {where_clause}"
        elif 'tat' in template_type.lower():
            # Include TAT_in_Days for TAT summaries
            query = f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State, TAT_in_Days FROM {table_name} {where_clause}"
        elif 'payreq' in template_type.lower():
            # Include Year_Pay_Req_Received for pay req summaries
            # Only select columns needed for summaries (grouping columns)
            # User should provide custom query if they need detail columns with different names
            query = f"SELECT Issue_State, Resident_State, Year_Pay_Req_Received FROM {table_name} {where_clause}"
        else:
            # Default query for state summaries or direct dump
            query = f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State FROM {table_name} {where_clause}"
    
    # Calculate detail columns count for spacing
    if detail_columns:
        detail_columns_count = len(detail_columns)
    elif template_type == 'state_summary_only':
        detail_columns_count = 0  # No detail columns
    else:
        # Estimate from query - count SELECT columns
        select_part = query.split('FROM')[0].replace('SELECT', '').strip()
        detail_columns_count = len([c.strip() for c in select_part.split(',')])
    
    # Create summary config based on template type
    summary_config = create_summary_config_from_template_type(template_type, detail_columns_count)
    
    # Determine spacing columns
    if template_type == 'state_summary_only':
        spacing_columns = []  # No detail, so no spacing needed
    elif summary_config is None:
        spacing_columns = []  # No summaries, so no spacing needed
    else:
        # Add one column gap between detail and summaries
        spacing_columns = [chr(ord('A') + detail_columns_count)]
    
    # Determine if filters should be enabled
    filters_enabled = template_type != 'state_summary_only'
    
    # Create formatting config with currency columns if provided
    formatting = FormattingConfig(
        header_row=1, 
        filters=filters_enabled,
        currency_columns=currency_columns
    )
    
    return WorksheetConfig(
        name=worksheet_name,
        query=query,
        detail_start_column='A',
        detail_columns=detail_columns,  # None means use actual column names from query
        spacing_columns=spacing_columns,
        summary_config=summary_config,
        formatting=formatting
    )


# ============================================================================
# Hardcoded Worksheet Definitions (Structure only - table names come from config)
# ============================================================================

def get_hardcoded_worksheet_structure(worksheet_name: str, table_name: str) -> Optional[WorksheetConfig]:
    """Return hardcoded worksheet structure based on template images. Table name is injected into query."""
    
    # Helper function to create standard worksheet configs
    def create_standard_worksheet(name, schedule_id, table_name, has_summary=True, highlight_cols=None):
        query = f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State FROM {table_name} WHERE Schedule_ID = '{schedule_id}'"
        detail_cols = ['Policy Num', 'Claim Num', 'Product', 'Claim Status', 'Company', 'Issue Sta', 'Resident Sta']
        formatting = FormattingConfig(header_row=1, filters=True, highlight_columns=highlight_cols or [])
        
        summary_configs = None
        if has_summary:
            summary_configs = [
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='I',
                    columns=['Issue State', 'CountOfPolicy No']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='K',
                    columns=['Resident State', 'CountOfPolicy No']
                )
            ]
        
        return WorksheetConfig(
            name=name,
            query=query,
            detail_start_column='A',
            detail_columns=detail_cols,
            spacing_columns=['H'],
            summary_config=summary_configs,
            formatting=formatting
        )
    
    # Summary Worksheet - handled separately, not returned here
    if worksheet_name == 'Summary':
        return None  # Summary is handled separately in create_workbook
    
    # Worksheet 1-001 - Summary only (no detail records)
    elif worksheet_name == '1-001':
        return WorksheetConfig(
            name='1-001',
            query=f"SELECT Policy_Num, Issue_State, Resident_State FROM {table_name} WHERE Schedule_ID = '1-001'",
            detail_start_column='A',
            detail_columns=None,  # None means use actual column names from query results
            spacing_columns=[],
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='A',  # Start from column A
                    columns=['Issue State', 'CountOfPolicy No']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='D',  # One column gap after Issue State summary (columns A-B, gap C, start D)
                    columns=['Resident State', 'CountOfPolicy No']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=False)  # No filters for summary-only
        )
    
    # Worksheet 1-004 - Detail records only (no summaries)
    elif worksheet_name == '1-004':
        return WorksheetConfig(
            name='1-004',
            query=f"SELECT Policy, Lapse_Da, Stati, Status_Reas, Company, Issue_St, Resident_St FROM {table_name} WHERE Schedule_ID = '1-004'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=[],
            summary_config=None,  # No summaries - detail only
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    # Worksheet 1-006 - Summary only (no detail records)
    elif worksheet_name == '1-006':
        return WorksheetConfig(
            name='1-006',
            query=f"SELECT Policy_Num, Issue_State, Resident_State FROM {table_name} WHERE Schedule_ID = '1-006'",
            detail_start_column='A',
            detail_columns=None,  # No detail columns - summary only
            spacing_columns=[],
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='A',  # Start from column A
                    columns=['Issue State', 'CountOfPolicy No']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='D',  # One column gap after Issue State summary (columns A-B, gap C, start D)
                    columns=['Resident State', 'CountOfPolicy No']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=False)  # No filters for summary-only
        )
    
    # Worksheet 2-001 - Detail records + Issue State/Resident State summaries
    elif worksheet_name == '2-001':
        return WorksheetConfig(
            name='2-001',
            query=f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State FROM {table_name} WHERE Schedule_ID = '2-001'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=['H'],  # One column gap between detail (A-G) and summaries
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='I',  # Start after gap column H (Issue State: I-J)
                    columns=['Issue State', 'Count']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='L',  # One column gap after Issue State summary (gap K, Resident State: L-M)
                    columns=['Resident State', 'Count']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    # Worksheet 2-002 - Same as 2-001
    elif worksheet_name == '2-002':
        return WorksheetConfig(
            name='2-002',
            query=f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State FROM {table_name} WHERE Schedule_ID = '2-002'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=['H'],  # One column gap between detail (A-G) and summaries
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='I',  # Issue State: I-J
                    columns=['Issue State', 'Count']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='L',  # Gap K, Resident State: L-M
                    columns=['Resident State', 'Count']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    # Worksheet 2-003 - Detail records + TAT summary by ranges with % of total
    elif worksheet_name == '2-003':
        return WorksheetConfig(
            name='2-003',
            query=f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State, TAT_in_Days FROM {table_name} WHERE Schedule_ID = '2-003'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=['I'],  # One column gap between detail (A-H) and TAT summary
            summary_config=[
                SummaryConfig(
                    group_by='TAT_Range',  # TAT summary by ranges (special handling)
                    aggregates=[AggregateConfig(field='TAT_in_Days', function='COUNT', label='TAT COUNTS')],
                    start_column='J',  # Start after gap column I (TAT Range: J-K-L)
                    columns=['', 'TAT COUNTS', '% of Total']  # Empty first column header (range values are in data rows)
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    # Worksheet 2-004 - Same as 2-001
    elif worksheet_name == '2-004':
        return WorksheetConfig(
            name='2-004',
            query=f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State FROM {table_name} WHERE Schedule_ID = '2-004'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=['H'],  # One column gap between detail (A-G) and summaries
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='I',  # Issue State: I-J
                    columns=['Issue State', 'Count']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='L',  # Gap K, Resident State: L-M
                    columns=['Resident State', 'Count']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    # Worksheet 2-005 - Same as 2-001
    elif worksheet_name == '2-005':
        return WorksheetConfig(
            name='2-005',
            query=f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State FROM {table_name} WHERE Schedule_ID = '2-005'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=['H'],  # One column gap between detail (A-G) and summaries
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='I',  # Issue State: I-J
                    columns=['Issue State', 'Count']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='L',  # Gap K, Resident State: L-M
                    columns=['Resident State', 'Count']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    # Worksheet 3-001 - Detail records + TAT summary (same pattern as 2-003)
    elif worksheet_name == '3-001':
        return WorksheetConfig(
            name='3-001',
            query=f"SELECT Inq_Date, Stat_Start_Date, Decision, Decision_Reason, Company, Issue_State, Resident_State, Product, Date_of_Loss, Schedule_ID, TAT_in_Days FROM {table_name} WHERE Schedule_ID = '3-001'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=['L'],  # One column gap between detail (A-K) and TAT summary
            summary_config=[
                SummaryConfig(
                    group_by='TAT_Range',  # TAT summary by ranges (same as 2-003)
                    aggregates=[AggregateConfig(field='TAT_in_Days', function='COUNT', label='TAT COUNTS')],
                    start_column='M',  # Start after gap column L (TAT Range: M-N-O)
                    columns=['', 'TAT COUNTS', '% of Total']  # Empty first column header
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True, highlight_columns=['C', 'D'])
        )
    
    # Worksheets 3-003, 3-004, 3-005 - Decision Details
    elif worksheet_name in ['3-003', '3-004', '3-005']:
        schedule_id = worksheet_name
        return WorksheetConfig(
            name=worksheet_name,
            query=f"SELECT Decision_Date, Inq_Date, Stat_Start_Date, Decision, Decision_Reason, Company, Issue_State, Resident_State, Product, Date_of_Loss, Schedule_ID, TAT_in_Days FROM {table_name} WHERE Schedule_ID = '{schedule_id}'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=[],
            summary_config=None,
            formatting=FormattingConfig(header_row=1, filters=True, highlight_columns=['D', 'E'])
        )
    
    # Worksheet 3-006
    elif worksheet_name == '3-006':
        return WorksheetConfig(
            name='3-006',
            query=f"SELECT Inq_Date, Stat_Start_Date, Decision, Decision_Reason, Company, Issue_State, Resident_State, Product, Date_of_Loss, Schedule_ID, TAT_in_Days FROM {table_name} WHERE Schedule_ID = '3-006'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=[],
            summary_config=None,
            formatting=FormattingConfig(header_row=1, filters=True, highlight_columns=['C', 'D'])
        )
    
    # Worksheet 3-007
    elif worksheet_name == '3-007':
        return WorksheetConfig(
            name='3-007',
            query=f"SELECT Policy, Claim_Number, Decision_Date, Inq_Date, Stat_Start_Date, Decision, Decision_Reason, Company, Issue_State, Resident_State, Product, Date_of_Loss, Schedule_ID_2 FROM {table_name} WHERE Schedule_ID = '3-007'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=[],
            summary_config=None,
            formatting=FormattingConfig(header_row=1, filters=True, highlight_columns=['F', 'G'])
        )
    
    # Worksheet 5-001 - Detail records + Issue State/Resident State summaries (like 2-002)
    elif worksheet_name == '5-001':
        return WorksheetConfig(
            name='5-001',
            query=f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State FROM {table_name} WHERE Schedule_ID = '5-001'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=['H'],  # One column gap between detail (A-G) and summaries
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='I',  # Issue State: I-J
                    columns=['Issue State', 'CountOfPolicy No']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='L',  # Gap K, Resident State: L-M
                    columns=['Resident State', 'CountOfPolicy No']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    # Worksheet 5-002 - Summary only (no detail records): Issue State/Resident State summaries with Company
    elif worksheet_name == '5-002':
        # Summary only: Issue State, Count, Company & Resident State, Count, Company
        return WorksheetConfig(
            name='5-002',
            query=f"SELECT Issue_State, Resident_State, Company, Count FROM {table_name} WHERE Schedule_ID = '5-002'",
            detail_start_column='A',
            detail_columns=None,  # No detail columns - summary only
            spacing_columns=[],  # No detail records, so no spacing needed
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[
                        AggregateConfig(field='Count', function='SUM', label='Count'),
                        AggregateConfig(field='Company', function='FIRST', label='Company')  # Display company name (first value)
                    ],
                    start_column='A',  # Issue State: A-B-C
                    columns=['Issue State', 'Count', 'Company']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[
                        AggregateConfig(field='Count', function='SUM', label='Count'),
                        AggregateConfig(field='Company', function='FIRST', label='Company')  # Display company name (first value)
                    ],
                    start_column='E',  # Gap D, Resident State: E-F-G
                    columns=['Resident State', 'Count', 'Company']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=False)  # No filters for summary-only
        )
    
    # Worksheet 5-003 - Detail records + TAT summary + Issue State/Resident State summaries
    elif worksheet_name == '5-003':
        return WorksheetConfig(
            name='5-003',
            query=f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State, TAT_in_Days FROM {table_name} WHERE Schedule_ID = '5-003'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=['I'],  # One column gap between detail (A-H) and summaries
            summary_config=[
                SummaryConfig(
                    group_by='TAT_Range',  # TAT summary by ranges (same as 2-003)
                    aggregates=[AggregateConfig(field='TAT_in_Days', function='COUNT', label='TAT COUNTS')],
                    start_column='J',  # TAT Range: J-K-L
                    columns=['', 'TAT COUNTS', '% of Total']
                ),
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='N',  # Gap M, Issue State: N-O
                    columns=['Issue State', 'CountOfPolicy No']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='Q',  # Gap P, Resident State: Q-R (one column gap after Issue State summary)
                    columns=['Resident State', 'CountOfPolicy No']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    # Worksheet 5-004 - Detail records + Issue State/Resident State summaries
    elif worksheet_name == '5-004':
        return WorksheetConfig(
            name='5-004',
            query=f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State FROM {table_name} WHERE Schedule_ID = '5-004'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=['H'],  # One column gap between detail (A-G) and summaries
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='I',  # Issue State: I-J
                    columns=['Issue State', 'CountOfPolicy No']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='Policy_Num', function='COUNT', label='CountOfPolicy No')],
                    start_column='L',  # Gap K, Resident State: L-M
                    columns=['Resident State', 'CountOfPolicy No']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    # Worksheet 6-001 - Summary only (no detail records): Like 5-002 (Issue State, Count, Company & Resident State, Count, Company)
    elif worksheet_name == '6-001':
        return WorksheetConfig(
            name='6-001',
            query=f"SELECT Issue_State, Resident_State, Company, Count FROM {table_name} WHERE Schedule_ID = '6-001'",
            detail_start_column='A',
            detail_columns=None,  # No detail columns - summary only
            spacing_columns=[],  # No detail records, so no spacing needed
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[
                        AggregateConfig(field='Count', function='SUM', label='Count'),
                        AggregateConfig(field='Company', function='FIRST', label='Company')  # Display company name (first value)
                    ],
                    start_column='A',  # Issue State: A-B-C
                    columns=['Issue State', 'Count', 'Company']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[
                        AggregateConfig(field='Count', function='SUM', label='Count'),
                        AggregateConfig(field='Company', function='FIRST', label='Company')  # Display company name (first value)
                    ],
                    start_column='E',  # Gap D, Resident State: E-F-G
                    columns=['Resident State', 'Count', 'Company']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=False)  # No filters for summary-only
        )
    
    # Worksheet 6-002 - Detail records + Issue State/Resident State summaries (Count, not CountOfPolicy No)
    elif worksheet_name == '6-002':
        return WorksheetConfig(
            name='6-002',
            query=f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State FROM {table_name} WHERE Schedule_ID = '6-002'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=['H'],  # One column gap between detail (A-G) and summaries
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='I',  # Issue State: I-J
                    columns=['Issue State', 'Count']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='L',  # Gap K, Resident State: L-M
                    columns=['Resident State', 'Count']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    # Worksheet 6-003 - Same as 6-002
    elif worksheet_name == '6-003':
        return WorksheetConfig(
            name='6-003',
            query=f"SELECT Policy_Num, Claim_Num, Product, Claim_Status, Company, Issue_State, Resident_State FROM {table_name} WHERE Schedule_ID = '6-003'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=['H'],  # One column gap between detail (A-G) and summaries
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='I',  # Issue State: I-J
                    columns=['Issue State', 'Count']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='L',  # Gap K, Resident State: L-M
                    columns=['Resident State', 'Count']
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    # Worksheet 6-004 - Same as 6-003 + Additional Summary (Counts, Year Pay Req Received)
    # NOTE: This hardcoded structure uses generic column names. For actual use, provide a custom query
    # with your actual column names (e.g., Policy instead of Policy_Num, doc_id, etc.)
    elif worksheet_name == '6-004':
        # Only select grouping columns to avoid "invalid identifier" errors
        # User should provide custom query with actual column names
        return WorksheetConfig(
            name='6-004',
            query=f"SELECT Issue_State, Resident_State, Year_Pay_Req_Received FROM {table_name} WHERE Schedule_ID = '6-004'",
            detail_start_column='A',
            detail_columns=None,  # Use actual column names from query results
            spacing_columns=[],  # No detail columns, so no spacing needed
            summary_config=[
                SummaryConfig(
                    group_by='Issue_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='A',  # Issue State: A-B
                    columns=['Issue State', 'Count']
                ),
                SummaryConfig(
                    group_by='Resident_State',
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Count')],  # Empty field = count records
                    start_column='D',  # Gap C, Resident State: D-E
                    columns=['Resident State', 'Count']
                ),
                SummaryConfig(
                    group_by='Year_Pay_Req_Received',  # Additional summary by Year
                    aggregates=[AggregateConfig(field='', function='COUNT', label='Counts')],  # Empty field = count records
                    start_column='G',  # Gap F, Year summary: G-H (Year Pay Req Received, Counts)
                    columns=['Year Pay Req Received', 'Counts']  # Year value first, then Counts
                )
            ],
            formatting=FormattingConfig(header_row=1, filters=True)
        )
    
    return None


# ============================================================================
# Snowflake Connection Functions
# ============================================================================

def create_snowflake_connection(config: SnowflakeConfig) -> snowflake.connector.SnowflakeConnection:
    """Create Snowflake connection with SSO or username/password"""
    try:
        conn_params = {
            'account': config.account,
            'warehouse': config.warehouse,
            'database': config.database,
            'schema': config.schema,
            'authenticator': config.authenticator
        }
        
        if config.authenticator == 'snowflake':
            if not config.user or not config.password:
                raise ValueError("User and password required for 'snowflake' authenticator")
            conn_params['user'] = config.user
            conn_params['password'] = config.password
        elif config.authenticator == 'externalbrowser':
            # For externalbrowser (SSO), user is optional
            # If provided and not empty, use it; otherwise don't pass it (Snowflake will use SSO user)
            if config.user and config.user.strip():
                conn_params['user'] = config.user.strip()
            # Note: Some Snowflake configurations may require user even for externalbrowser
            # If you get "user is empty" error, set SNOWFLAKE_USER environment variable
        
        connection = snowflake.connector.connect(**conn_params)
        print(f"Successfully connected to Snowflake account: {config.account}")
        return connection
    except Exception as e:
        error_msg = str(e)
        if 'user' in error_msg.lower() and 'empty' in error_msg.lower():
            print(f"Error connecting to Snowflake: {e}")
            print("\nTROUBLESHOOTING:")
            print("Your Snowflake configuration requires a USER parameter even for SSO authentication.")
            print("Please set the SNOWFLAKE_USER environment variable:")
            print("  PowerShell: $env:SNOWFLAKE_USER = 'your_username'")
            print("  Command Prompt: set SNOWFLAKE_USER=your_username")
            print("  Or add 'user: your_username' to the snowflake section in config.yaml")
        else:
            print(f"Error connecting to Snowflake: {e}")
        sys.exit(1)


def resolve_table_names_in_query(query: str, database: str, schema: str) -> str:
    """
    Resolve unqualified table names in SQL query to database.schema.table_name format.
    
    Args:
        query: SQL query string
        database: Database name to prepend
        schema: Schema name to prepend
    
    Returns:
        Query with resolved table names
    """
    if not database or not schema:
        return query  # Can't resolve without database/schema
    
    # Reserved SQL keywords that should not be treated as table names
    reserved_words = {'SELECT', 'FROM', 'WHERE', 'JOIN', 'INNER', 'LEFT', 'RIGHT', 'FULL', 
                     'OUTER', 'ON', 'ORDER', 'GROUP', 'BY', 'HAVING', 'UNION', 'INTERSECT', 
                     'EXCEPT', 'UPDATE', 'INSERT', 'INTO', 'DELETE', 'CREATE', 'ALTER', 'DROP',
                     'TABLE', 'VIEW', 'INDEX', 'DATABASE', 'SCHEMA', 'AS', 'AND', 'OR', 'NOT',
                     'IN', 'EXISTS', 'LIKE', 'BETWEEN', 'IS', 'NULL', 'TRUE', 'FALSE'}
    
    # Pattern to match table names after FROM, JOIN, UPDATE, INSERT INTO, etc.
    # This pattern matches unqualified identifiers (no dots, not quoted)
    def replace_table_in_context(match):
        full_match = match.group(0)
        table_name = match.group(1)
        
        # Skip if already qualified (contains dot) - means it's schema.table or database.schema.table
        if '.' in table_name:
            return full_match
        
        # Skip if it's a reserved word
        if table_name.upper() in reserved_words:
            return full_match
        
        # Check if it's quoted (skip quoted identifiers - though pattern shouldn't match these)
        if table_name.startswith('"') or table_name.startswith("'") or table_name.startswith('`'):
            return full_match
        
        # Resolve to database.schema.table_name
        return full_match.replace(table_name, f"{database}.{schema}.{table_name}", 1)
    
    # Patterns for different SQL contexts
    # FROM table_name [alias]
    query = re.sub(r'\bFROM\s+([a-zA-Z_][a-zA-Z0-9_]*(?:\.[a-zA-Z_][a-zA-Z0-9_]*)*)(?:\s+[a-zA-Z_][a-zA-Z0-9_]*)?\b', 
                  replace_table_in_context, query, flags=re.IGNORECASE)
    
    # JOIN table_name [alias] (handles INNER JOIN, LEFT JOIN, etc.)
    query = re.sub(r'\b(?:INNER|LEFT|RIGHT|FULL|CROSS)?\s*JOIN\s+([a-zA-Z_][a-zA-Z0-9_]*(?:\.[a-zA-Z_][a-zA-Z0-9_]*)*)(?:\s+[a-zA-Z_][a-zA-Z0-9_]*)?\b',
                  replace_table_in_context, query, flags=re.IGNORECASE)
    
    # UPDATE table_name
    query = re.sub(r'\bUPDATE\s+([a-zA-Z_][a-zA-Z0-9_]*(?:\.[a-zA-Z_][a-zA-Z0-9_]*)*)\b',
                  replace_table_in_context, query, flags=re.IGNORECASE)
    
    # INSERT INTO table_name
    query = re.sub(r'\bINSERT\s+INTO\s+([a-zA-Z_][a-zA-Z0-9_]*(?:\.[a-zA-Z_][a-zA-Z0-9_]*)*)\b',
                  replace_table_in_context, query, flags=re.IGNORECASE)
    
    return query


def execute_query(connection: snowflake.connector.SnowflakeConnection, query: str, 
                 database: Optional[str] = None, schema: Optional[str] = None) -> List[Dict[str, Any]]:
    """Execute SQL query and return results as list of dictionaries"""
    original_query = query
    try:
        # Resolve table names if database and schema are provided
        if database and schema:
            query = resolve_table_names_in_query(query, database, schema)
        
        # Debug: Print the query being executed (first 500 chars)
        print(f"  DEBUG: Executing query (first 500 chars): {query[:500]}")
        if len(query) > 500:
            print(f"  DEBUG: ... (query continues, total length: {len(query)} chars)")
        
        # Use a separate cursor for each query (thread-safe)
        cursor = connection.cursor()
        try:
            cursor.execute(query)
            
            # Get column names
            columns = [desc[0] for desc in cursor.description]
            
            # Fetch all rows
            rows = cursor.fetchall()
            
            # Convert to list of dictionaries
            results = []
            for row in rows:
                row_dict = {}
                for i, col in enumerate(columns):
                    row_dict[col] = row[i]
                results.append(row_dict)
            
            return results
        finally:
            cursor.close()
    except Exception as e:
        error_msg = str(e)
        print(f"Error executing query: {error_msg}")
        print(f"\nOriginal query:")
        print(f"  {original_query[:500]}")
        if query != original_query:
            print(f"\nResolved query (after table name resolution):")
            print(f"  {query[:500]}")
        
        # Provide helpful error messages for common issues
        if 'invalid identifier' in error_msg.lower() or 'does not exist' in error_msg.lower():
            print("\nTROUBLESHOOTING:")
            print("Common causes:")
            print("1. Column name doesn't exist - check spelling and case")
            print("2. Wrong syntax in WHERE clause:")
            print("   ❌ WHERE 'Company' = 'CCC'  (wrong - 'Company' is treated as string)")
            print("   ✅ WHERE Company = 'CCC'     (correct - Company is column, 'CCC' is value)")
            print("3. Table name not found - check if table exists in database.schema")
            print("4. If using quoted identifiers, they must match exact case")
        
        raise


def close_connection(connection: snowflake.connector.SnowflakeConnection):
    """Close Snowflake connection"""
    try:
        connection.close()
        print("Snowflake connection closed")
    except Exception as e:
        print(f"Error closing connection: {e}")


# ============================================================================
# Data Processing Functions
# ============================================================================

def fetch_detail_records(connection: snowflake.connector.SnowflakeConnection, query: str,
                        database: Optional[str] = None, schema: Optional[str] = None) -> List[Dict[str, Any]]:
    """Fetch detail records from Snowflake"""
    return execute_query(connection, query, database, schema)


def get_tat_range(tat_value: Any) -> str:
    """
    Categorize TAT value into range buckets based on TAT_in_Days column.
    
    Categories:
    - -1 < x < 31: 0 to 30 days
    - x > 30 and x < 61: 31 to 60 days
    - 60 to 91: 61 to 90 days
    - > 90: 91+ days
    """
    try:
        if tat_value is None:
            return None
        tat = float(tat_value)
        if -1 < tat < 31:  # -1 < x < 31 (i.e., 0 to 30)
            return "-1 to <31"
        elif 30 < tat < 61:  # x > 30 and x < 61 (i.e., 31 to 60)
            return ">30 and <61"
        elif 60 < tat <= 90:  # 60 to 91 (i.e., 61 to 90)
            return ">60 and <91"
        elif tat > 90:  # > 90 (i.e., 91 and above)
            return ">90"
        else:
            return None  # Exclude from summary (handles edge cases like exactly -1, 30, 60, 91)
    except (ValueError, TypeError):
        return None  # Exclude from summary


def generate_summary(detail_records: List[Dict[str, Any]], summary_config: SummaryConfig, include_grand_total: bool = True) -> List[Dict[str, Any]]:
    """Generate summary table from detail records with optional Grand Total row and percentage calculation"""
    # Group records by the specified field
    grouped = defaultdict(list)
    
    # Special handling for TAT_Range grouping
    if summary_config.group_by == 'TAT_Range':
        # Helper for case-insensitive field access
        def get_tat_value_case_insensitive(record, field_name):
            """Get TAT_in_Days value with case-insensitive and space-insensitive matching"""
            # Try exact match first
            if field_name in record:
                return record[field_name]
            # Try case-insensitive match
            for key, value in record.items():
                if key.upper() == field_name.upper():
                    return value
                # Also try matching with spaces normalized (replace spaces/underscores)
                normalized_key = key.replace(' ', '_').replace('-', '_').upper()
                normalized_field = field_name.replace(' ', '_').replace('-', '_').upper()
                if normalized_key == normalized_field:
                    return value
            # Try common variations (with underscores, spaces, and hyphens)
            variations = [
                'TAT_IN_DAYS', 'tat_in_days', 'Tat_In_Days',
                'TAT IN DAYS', 'Tat In Days', 'tat in days',
                'TAT-IN-DAYS', 'Tat-In-Days', 'tat-in-days',
                'TAT_DAYS', 'tat_days', 'Tat_Days',
                'TAT DAYS', 'Tat Days', 'tat days',
                'TAT', 'tat', 'Tat'
            ]
            for var in variations:
                if var in record:
                    return record[var]
            return None
        
        # Debug: Check if TAT column exists in data
        if detail_records:
            sample_record = detail_records[0]
            tat_found = False
            tat_column_name = None
            for key in sample_record.keys():
                normalized_key = key.replace(' ', '_').replace('-', '_').upper()
                if 'TAT' in normalized_key and 'DAY' in normalized_key:
                    tat_found = True
                    tat_column_name = key
                    break
            
            if not tat_found:
                print(f"  WARNING: TAT column not found in data. Available columns: {list(sample_record.keys())}")
                print(f"  Looking for columns containing 'TAT' and 'DAY' (case-insensitive, spaces/underscores ignored)")
            else:
                print(f"  DEBUG: Found TAT column: '{tat_column_name}'")
        
        records_with_tat = 0
        records_without_tat = 0
        records_by_range = {}
        
        for record in detail_records:
            # Get TAT_in_Days value with case-insensitive matching and categorize into range
            tat_value = get_tat_value_case_insensitive(record, 'TAT_in_Days')
            if tat_value is None:
                records_without_tat += 1
                continue
            
            records_with_tat += 1
            group_key = get_tat_range(tat_value)
            # Include all records with valid TAT ranges
            if group_key is not None:
                grouped[group_key].append(record)
                records_by_range[group_key] = records_by_range.get(group_key, 0) + 1
            else:
                # Debug: show records that were excluded (only first few to avoid spam)
                if len([r for r in detail_records if get_tat_value_case_insensitive(r, 'TAT_in_Days') == tat_value]) <= 3:
                    print(f"  DEBUG: Excluded record with TAT_in_Days={tat_value} (not in valid range)")
        
        # Debug output for TAT summary
        print(f"  DEBUG: TAT Summary - Total records: {len(detail_records)}, With TAT: {records_with_tat}, Without TAT: {records_without_tat}")
        print(f"  DEBUG: Records by TAT range: {records_by_range}")
        print(f"  DEBUG: Grouped records count: {sum(len(v) for v in grouped.values())}")
    else:
        # Standard grouping by field value
        # Handle case-insensitive column name matching
        def get_field_value_case_insensitive(record, field_name):
            """Get field value from record with case-insensitive matching"""
            # Try exact match first
            if field_name in record:
                return record[field_name]
            # Try case-insensitive match
            for key, value in record.items():
                if key.upper() == field_name.upper():
                    return value
            # Try common variations
            variations = [
                field_name.replace('_', ''),
                field_name.lower(),
                field_name.upper(),
                field_name.replace('_', ' '),
                # Handle shortened versions (e.g., Resident_St vs Resident_State)
                field_name.replace('_State', '_St'),
                field_name.replace('_State', 'State'),
                field_name.replace('State', 'St'),
                # Handle Issue_State variations
                field_name.replace('Issue_State', 'Issue_St'),
                field_name.replace('Issue_State', 'IssueState'),
                # Handle Resident_State variations
                field_name.replace('Resident_State', 'Resident_St'),
                field_name.replace('Resident_State', 'ResidentState'),
            ]
            for var in variations:
                if var in record:
                    return record[var]
            # Try partial matching (e.g., if looking for "Resident_State" but column is "Resident_St")
            field_upper = field_name.upper()
            for key in record.keys():
                key_upper = key.upper()
                # Check if field_name is a prefix of key or vice versa (for shortened names)
                if (field_upper.startswith(key_upper) or key_upper.startswith(field_upper)) and len(key_upper) >= len(field_upper) * 0.7:
                    return record[key]
            return None
        
        for record in detail_records:
            group_key = get_field_value_case_insensitive(record, summary_config.group_by)
            if group_key is None:
                group_key = ''
            # Convert to string and handle None/empty values
            group_key = str(group_key) if group_key is not None else ''
            grouped[group_key].append(record)
    
    # Calculate aggregates for each group
    summary_rows = []
    grand_total_values = {}
    total_count = 0  # For percentage calculation
    
    # Define TAT range order for proper sorting
    tat_range_order = ["-1 to <31", ">30 and <61", ">60 and <91", ">90"]
    
    # Sort groups - use TAT range order if grouping by TAT_Range, otherwise alphabetical
    if summary_config.group_by == 'TAT_Range':
        sorted_groups = sorted(grouped.items(), key=lambda x: (
            tat_range_order.index(x[0]) if x[0] in tat_range_order else len(tat_range_order)
        ))
    else:
        sorted_groups = sorted(grouped.items())
    
    # Debug output for grouping issues
    if len(grouped) == 1:
        first_key = list(grouped.keys())[0]
        if first_key in ('', 'None', None):
            print(f"  WARNING: All {len(detail_records)} records grouped into single group (empty/None values)")
            print(f"  Group by field: '{summary_config.group_by}'")
            if detail_records:
                print(f"  Available columns in data: {list(detail_records[0].keys())}")
                # Try to find the field
                sample_record = detail_records[0]
                found_value = None
                if summary_config.group_by in sample_record:
                    found_value = sample_record[summary_config.group_by]
                else:
                    for key, val in sample_record.items():
                        if key.upper() == summary_config.group_by.upper():
                            found_value = val
                            break
                print(f"  Sample '{summary_config.group_by}' value from first record: {found_value}")
                print(f"  This suggests the column name doesn't match. Check your query column names.")
    
    # First pass: calculate counts and totals
    for group_key, records in sorted_groups:
        # For TAT_Range grouping, use empty string for first column (users know row position = category)
        if summary_config.group_by == 'TAT_Range':
            summary_row = {summary_config.columns[0]: ''}  # Empty first column - users know row position
        else:
            summary_row = {summary_config.columns[0]: group_key}
        
        # Debug output for state_summary_with_company template
        if len(summary_config.aggregates) > 1 and any(agg.function.upper() in ('SUM', 'FIRST') for agg in summary_config.aggregates):
            if records:
                print(f"  DEBUG: Group '{group_key}' has {len(records)} records")
                print(f"  DEBUG: First record columns: {list(records[0].keys())}")
                for agg in summary_config.aggregates:
                    print(f"  DEBUG: Looking for field '{agg.field}' with function '{agg.function}'")
        
        for agg in summary_config.aggregates:
            field_value = agg.field
            function = agg.function.upper()
            
            if function == 'COUNT':
                # COUNT just counts records - field name is not used, but we can use it for distinct counting if needed
                # If field is specified and not empty, try to count distinct values
                if field_value and field_value.strip():
                    # Helper for case-insensitive field access
                    def get_field_val(record, field_name):
                        if field_name in record:
                            return record[field_name]
                        for key, val in record.items():
                            if key.upper() == field_name.upper():
                                return val
                        # Try variations
                        variations = [
                            field_name.replace('_', ''),
                            field_name.lower(),
                            field_name.upper(),
                            field_name.replace('_', ' '),
                        ]
                        for var in variations:
                            if var in record:
                                return record[var]
                        return None
                    # Count distinct values of the field
                    unique_values = set()
                    for record in records:
                        val = get_field_val(record, field_value)
                        if val is not None and val != '':
                            unique_values.add(val)
                    value = len(unique_values) if unique_values else len(records)
                else:
                    # No field specified - just count records
                    value = len(records)
                total_count += value  # Accumulate total for percentage
            elif function == 'SUM':
                # Helper for case-insensitive field access
                def get_field_val(record, field_name):
                    if field_name in record:
                        return record[field_name]
                    for key, val in record.items():
                        if key.upper() == field_name.upper():
                            return val
                    # Try variations
                    variations = [
                        field_name.replace('_', ''),
                        field_name.lower(),
                        field_name.upper(),
                        field_name.replace('_', ' '),
                    ]
                    for var in variations:
                        if var in record:
                            return record[var]
                    return None
                # Sum values, handling both numeric strings and numbers
                sum_values = []
                for r in records:
                    val = get_field_val(r, field_value)
                    if val is not None and val != '':
                        try:
                            sum_values.append(float(val))
                        except (ValueError, TypeError):
                            # If conversion fails, try to extract number from string
                            if isinstance(val, str):
                                # Try to extract first number from string
                                match = re.search(r'\d+\.?\d*', str(val))
                                if match:
                                    sum_values.append(float(match.group()))
                value = sum(sum_values) if sum_values else 0
            elif function == 'AVG' or function == 'AVERAGE':
                def get_field_val(record, field_name):
                    if field_name in record:
                        return record[field_name]
                    for key, val in record.items():
                        if key.upper() == field_name.upper():
                            return val
                    return None
                values = [float(get_field_val(r, field_value) or 0) for r in records if get_field_val(r, field_value) is not None]
                value = sum(values) / len(values) if values else 0
            elif function == 'MIN':
                def get_field_val(record, field_name):
                    if field_name in record:
                        return record[field_name]
                    for key, val in record.items():
                        if key.upper() == field_name.upper():
                            return val
                    return None
                values = [get_field_val(r, field_value) for r in records if get_field_val(r, field_value) is not None]
                value = min(values) if values else None
            elif function == 'MAX':
                def get_field_val(record, field_name):
                    if field_name in record:
                        return record[field_name]
                    for key, val in record.items():
                        if key.upper() == field_name.upper():
                            return val
                    return None
                values = [get_field_val(r, field_value) for r in records if get_field_val(r, field_value) is not None]
                value = max(values) if values else None
            elif function == 'FIRST' or function == 'VALUE':
                # Get first non-null value from the field (for displaying company name, etc.)
                # Helper for case-insensitive field access
                def get_field_val(record, field_name):
                    if field_name in record:
                        return record[field_name]
                    for key, val in record.items():
                        if key.upper() == field_name.upper():
                            return val
                    # Try variations
                    variations = [
                        field_name.replace('_', ''),
                        field_name.lower(),
                        field_name.upper(),
                        field_name.replace('_', ' '),
                    ]
                    for var in variations:
                        if var in record:
                            return record[var]
                    return None
                value = ''
                for record in records:
                    val = get_field_val(record, field_value)
                    if val is not None and val != '':
                        value = val
                        break
            else:
                value = 0
            
            summary_row[agg.label] = value
            
            # Debug output for state_summary_with_company
            if len(summary_config.aggregates) > 1 and any(a.function.upper() in ('SUM', 'FIRST') for a in summary_config.aggregates):
                print(f"  DEBUG: Aggregate '{agg.label}' (field='{field_value}', function='{function}') = {value}")
            
            # Accumulate for grand total
            if include_grand_total:
                if agg.label not in grand_total_values:
                    grand_total_values[agg.label] = 0
                if isinstance(value, (int, float)):
                    grand_total_values[agg.label] += value
        
        summary_rows.append(summary_row)
    
    # Second pass: calculate percentages if '% of Total' column exists
    if '% of Total' in summary_config.columns and total_count > 0:
        for summary_row in summary_rows:
            # Find the count value (should be the aggregate label)
            count_value = None
            for agg in summary_config.aggregates:
                if agg.label in summary_row:
                    count_value = summary_row[agg.label]
                    break
            
            if count_value is not None and isinstance(count_value, (int, float)):
                percentage = (count_value / total_count) * 100
                # Format with % symbol
                summary_row['% of Total'] = f"{round(percentage, 2)}%"
            else:
                summary_row['% of Total'] = "0%"
    
    # Add Grand Total row if requested
    if include_grand_total:
        # For TAT_Range, first column should be empty (no "Grand Total" label)
        if summary_config.group_by == 'TAT_Range':
            grand_total_row = {summary_config.columns[0]: ''}  # Empty for TAT range
        else:
            grand_total_row = {summary_config.columns[0]: 'Grand Total'}
        for agg in summary_config.aggregates:
            grand_total_row[agg.label] = grand_total_values.get(agg.label, 0)
        
        # Add percentage for grand total (should be 100%)
        if '% of Total' in summary_config.columns:
            grand_total_row['% of Total'] = "100%"
        
        summary_rows.append(grand_total_row)
    
    return summary_rows


def process_worksheet_data(connection: snowflake.connector.SnowflakeConnection, 
                           worksheet_config: WorksheetConfig,
                           database: Optional[str] = None, schema: Optional[str] = None) -> Tuple[List[Dict[str, Any]], List[List[Dict[str, Any]]]]:
    """Process worksheet data: fetch details and generate summaries"""
    # Fetch detail records
    detail_records = fetch_detail_records(connection, worksheet_config.query, database, schema)
    
    # Generate summaries if configured
    # Parallelize summary generation if there are multiple summaries (CPU-bound operation)
    summaries = []
    if worksheet_config.summary_config:
        if len(worksheet_config.summary_config) > 1:
            # Multiple summaries - generate in parallel for better performance
            def generate_single_summary(sum_config):
                include_grand_total = worksheet_config.name in ['1-001', '1-004', '1-006']
                if sum_config.group_by in ['Issue_State', 'Resident_State']:
                    include_grand_total = True
                return generate_summary(detail_records, sum_config, include_grand_total=include_grand_total)
            
            with ThreadPoolExecutor(max_workers=len(worksheet_config.summary_config)) as executor:
                # Create a mapping of index to sum_config for proper ordering
                summary_futures = {}
                for idx, sum_config in enumerate(worksheet_config.summary_config):
                    future = executor.submit(generate_single_summary, sum_config)
                    summary_futures[future] = (idx, sum_config)
                
                # Collect results
                summary_results = {}
                for future in as_completed(summary_futures):
                    idx, sum_config = summary_futures[future]
                    try:
                        summary_data = future.result()
                        summary_results[idx] = summary_data
                    except Exception as e:
                        print(f"  Error generating summary for '{sum_config.group_by}': {e}")
                        summary_results[idx] = []
                
                # Reorder summaries to match original config order
                for idx in range(len(worksheet_config.summary_config)):
                    summaries.append(summary_results.get(idx, []))
        else:
            # Single summary - generate sequentially
            for sum_config in worksheet_config.summary_config:
                # Include Grand Total for Schedule 1 worksheets (1-001, 1-004, 1-006)
                # Also include for all state summaries (Issue_State, Resident_State)
                include_grand_total = worksheet_config.name in ['1-001', '1-004', '1-006']
                # For state summaries (Issue_State, Resident_State), always include grand total
                if sum_config.group_by in ['Issue_State', 'Resident_State']:
                    include_grand_total = True
                print(f"  DEBUG: Generating summary for '{sum_config.group_by}', include_grand_total={include_grand_total}")
                summary_data = generate_summary(detail_records, sum_config, include_grand_total=include_grand_total)
                # Debug: Check if grand total row was added
                if summary_data:
                    has_grand_total = any(row.get(sum_config.columns[0]) == 'Grand Total' for row in summary_data)
                    print(f"  DEBUG: Summary has {len(summary_data)} rows, grand total present: {has_grand_total}")
                summaries.append(summary_data)
    
    return detail_records, summaries


# ============================================================================
# Excel Generation Functions
# ============================================================================

def column_letter_to_index(column_letter: str) -> int:
    """Convert column letter (e.g., 'A', 'Z', 'AA') to 1-based index"""
    result = 0
    for char in column_letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def index_to_column_letter(index: int) -> str:
    """Convert 1-based index to column letter"""
    return get_column_letter(index)


def apply_cell_formatting(ws, row: int, col: int, value: Any, is_header: bool = False):
    """Apply formatting to a cell"""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    
    if is_header:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Apply Orange Accent 2 Light 60% fill (approximately #FFD966)
        orange_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        cell.fill = orange_fill
    else:
        cell.alignment = Alignment(vertical='top')
    
    # Apply borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def apply_highlight(ws, row: int, col: int):
    """Apply yellow highlight to a cell"""
    cell = ws.cell(row=row, column=col)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    cell.fill = yellow_fill

def apply_border(ws, row: int, col: int):
    """Apply thin border to a cell"""
    cell = ws.cell(row=row, column=col)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def write_detail_table(ws, detail_records: List[Dict[str, Any]], 
                      worksheet_config: WorksheetConfig, start_row: int = 1):
    """Write detail table to worksheet - always writes headers even if no data"""
    # Determine column names from config or use empty list if no records
    if detail_records:
        # Get column names from first record or use configured columns
        if worksheet_config.detail_columns and len(worksheet_config.detail_columns) > 0:
            # Use configured display column names
            columns = worksheet_config.detail_columns
            # Map display names to actual column names from query results
            actual_columns = list(detail_records[0].keys())
            column_map = {}
            for i, display_col in enumerate(columns):
                if i < len(actual_columns):
                    # Try to match by position first
                    column_map[display_col] = actual_columns[i]
                else:
                    # If no match by position, try to find by name (case-insensitive)
                    found = False
                    for actual_col in actual_columns:
                        if actual_col.lower().replace('_', ' ') == display_col.lower():
                            column_map[display_col] = actual_col
                            found = True
                            break
                    if not found:
                        column_map[display_col] = display_col
        else:
            # Use actual column names from query results
            columns = list(detail_records[0].keys())
            column_map = {col: col for col in columns}
    else:
        # No data - use configured columns if available, otherwise we'll write empty headers
        if worksheet_config.detail_columns and len(worksheet_config.detail_columns) > 0:
            columns = worksheet_config.detail_columns
            column_map = {col: col for col in columns}
        else:
            # No data and no configured columns - can't determine headers
            print(f"  WARNING: No data and no detail_columns configured for worksheet '{worksheet_config.name}'. Cannot write headers.")
            return start_row
    
    # Write headers (always write headers even if no data)
    start_col_idx = column_letter_to_index(worksheet_config.detail_start_column)
    header_row = worksheet_config.formatting.header_row if worksheet_config.formatting else start_row
    
    for col_idx, col_name in enumerate(columns):
        apply_cell_formatting(ws, header_row, start_col_idx + col_idx, col_name, is_header=True)
    
    # If no data, return after writing headers
    if not detail_records:
        print(f"  INFO: No data for worksheet '{worksheet_config.name}', but headers have been written.")
        return header_row + 1  # Return next row after header
    
    # Write data rows
    data_start_row = header_row + 1
    for row_idx, record in enumerate(detail_records):
        for col_idx, col_name in enumerate(columns):
            actual_col = column_map.get(col_name, col_name)
            value = record.get(actual_col, '')
            cell = ws.cell(row=data_start_row + row_idx, column=start_col_idx + col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical='top')
            apply_border(ws, data_start_row + row_idx, start_col_idx + col_idx)
            
            # Apply currency formatting if configured
            if worksheet_config.formatting and worksheet_config.formatting.currency_columns:
                # Check if this column (by name) should be formatted as currency
                # Support case-insensitive matching
                should_format_currency = False
                for currency_col in worksheet_config.formatting.currency_columns:
                    if col_name.upper() == currency_col.upper() or actual_col.upper() == currency_col.upper():
                        should_format_currency = True
                        break
                
                if should_format_currency and value is not None and value != '':
                    try:
                        # Convert to float if it's a number
                        numeric_value = float(value)
                        # Apply currency format: $#,##0.00
                        cell.number_format = '$#,##0.00'
                        cell.value = numeric_value
                    except (ValueError, TypeError):
                        # If conversion fails, leave as is (might be a string)
                        pass
            
            # Apply highlighting if configured
            col_letter = index_to_column_letter(start_col_idx + col_idx)
            if worksheet_config.formatting:
                if col_letter in (worksheet_config.formatting.highlight_columns or []):
                    apply_highlight(ws, data_start_row + row_idx, start_col_idx + col_idx)
    
    # Apply filters if configured
    if worksheet_config.formatting and worksheet_config.formatting.filters:
        end_col_idx = start_col_idx + len(columns) - 1
        ws.auto_filter.ref = f"{index_to_column_letter(start_col_idx)}{header_row}:{index_to_column_letter(end_col_idx)}{data_start_row + len(detail_records)}"
    
    return data_start_row + len(detail_records)


def write_summary_table(ws, summary_data: List[Dict[str, Any]], 
                        summary_config: SummaryConfig, start_row: int = 1):
    """Write summary table to worksheet with Grand Total row formatting"""
    if not summary_data:
        print(f"  WARNING: No summary data to write for group_by='{summary_config.group_by}'. Summary table will be empty.")
        return
    
    print(f"  DEBUG: Writing summary table for group_by='{summary_config.group_by}' with {len(summary_data)} rows")
    
    start_col_idx = column_letter_to_index(summary_config.start_column)
    header_row = start_row
    
    # Write headers
    for col_idx, col_name in enumerate(summary_config.columns):
        apply_cell_formatting(ws, header_row, start_col_idx + col_idx, col_name, is_header=True)
    
    # Write data rows
    data_start_row = header_row + 1
    for row_idx, summary_row in enumerate(summary_data):
        is_grand_total = summary_row.get(summary_config.columns[0]) == 'Grand Total'
        
        for col_idx, col_name in enumerate(summary_config.columns):
            value = summary_row.get(col_name, '')
            cell = ws.cell(row=data_start_row + row_idx, column=start_col_idx + col_idx)
            cell.value = value
            
            # Format Grand Total row (bold, but not orange header)
            if is_grand_total:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='right' if col_idx > 0 else 'left', vertical='top')
            else:
                cell.alignment = Alignment(horizontal='right' if col_idx > 0 else 'left', vertical='top')
            
            # Apply borders
            apply_border(ws, data_start_row + row_idx, start_col_idx + col_idx)
        
        # Add Grand Total value in the next cell after CountOfPolicy No (if this is the last summary column)
        if is_grand_total and col_idx == len(summary_config.columns) - 1:
            # The Grand Total value is already in the CountOfPolicy No column
            # But if we need to add it in a separate cell, we can do it here
            pass
    
    return data_start_row + len(summary_data)


def format_date_for_reporting_period(date_str: str) -> str:
    """
    Format date string into format like "January 1, 2024"
    Accepts formats: YYYY-MM-DD, MM/DD/YYYY, DD-MM-YYYY, etc.
    """
    from datetime import datetime
    
    # Try common date formats
    date_formats = [
        '%Y-%m-%d',      # 2024-01-01
        '%m/%d/%Y',      # 01/01/2024
        '%d/%m/%Y',      # 01/01/2024 (European)
        '%Y/%m/%d',      # 2024/01/01
        '%m-%d-%Y',      # 01-01-2024
        '%d-%m-%Y',      # 01-01-2024 (European)
        '%B %d, %Y',     # January 1, 2024
        '%b %d, %Y',     # Jan 1, 2024
    ]
    
    for fmt in date_formats:
        try:
            dt = datetime.strptime(date_str.strip(), fmt)
            # Format as "January 1, 2024" (remove leading zero from day)
            day = dt.day
            month_name = dt.strftime('%B')
            year = dt.year
            return f"{month_name} {day}, {year}"
        except ValueError:
            continue
    
    # If no format matches, return as-is
    return date_str


def create_summary_worksheet(wb, summary_data, schedule_titles=None, reporting_period=None):
    """
    Create Summary worksheet with all 6 schedules, proper spacing
    summary_data: List of dicts with keys: Schedule_ID, Description, Value
    schedule_titles: Dict mapping schedule number to title
    reporting_period: String like "January 1, 2024 through December 31, 2024" or None to use default
    """
    # Validate input
    if not summary_data:
        print("  WARNING: No summary data provided to create_summary_worksheet!")
        summary_data = []
    
    ws = wb.create_sheet(title="Summary")
    
    # Font: Aptos Narrow, size 12
    default_font = Font(name='Aptos Narrow', size=12)
    bold_font = Font(name='Aptos Narrow', size=12, bold=True)
    
    # Default schedule titles if not provided
    if schedule_titles is None:
        schedule_titles = {
            1: "Schedule 1 - General Information",
            2: "Schedule 2 - Claimants",
            3: "Schedule 3 - Claimant Requests Denied/Not Paid",
            4: "Schedule 4",
            5: "Schedule 5",
            6: "Schedule 6"
        }
    
    # Default reporting period if not provided
    if reporting_period is None:
        reporting_period = "January 1, 2024 through December 31, 2024"
    
    # Leave first row empty (row 1)
    # Header Section - Start from row 2
    ws['A2'] = "Line of Business: Individual Long-Term Care"
    ws['A3'] = f"Reporting Period: {reporting_period}"
    ws['A4'] = "Filing Deadline: n/a"
    
    # Apply fonts and alignment (left aligned)
    ws['A2'].font = bold_font
    ws['A2'].alignment = Alignment(horizontal='left', vertical='top')
    
    ws['A3'].font = bold_font
    ws['A3'].alignment = Alignment(horizontal='left', vertical='top')
    apply_highlight(ws, 3, 1)  # Yellow highlight
    
    ws['A4'].font = bold_font
    ws['A4'].alignment = Alignment(horizontal='left', vertical='top')
    apply_highlight(ws, 4, 1)  # Yellow highlight
    
    # Group data by schedule number (1-xxx, 2-xxx, 3-xxx, etc.)
    schedule_groups = defaultdict(list)
    
    # Handle case-insensitive column name matching
    def get_value_case_insensitive(row, key):
        """Get value from row with case-insensitive key matching"""
        for k, v in row.items():
            if k.upper() == key.upper():
                return v
        return row.get(key, '')
    
    for row in summary_data:
        # Try to get Schedule_ID with case-insensitive matching
        # Accept: Schedule_ID, SCHEDULE_ID, schedule_id, ScheduleID, SCHEDULEID, ID, id
        schedule_id = get_value_case_insensitive(row, 'Schedule_ID') or \
                     get_value_case_insensitive(row, 'ScheduleID') or \
                     get_value_case_insensitive(row, 'SCHEDULEID') or \
                     get_value_case_insensitive(row, 'ID') or \
                     str(row.get(list(row.keys())[0], '')) if row else ''
        
        # Group by schedule number if Schedule_ID contains a dash (e.g., "1-001", "2-003")
        if '-' in str(schedule_id):
            try:
                schedule_num = int(str(schedule_id).split('-')[0])
                schedule_groups[schedule_num].append(row)
            except (ValueError, IndexError):
                # If parsing fails, put in a default group (0)
                schedule_groups[0].append(row)
        else:
            # If no dash, try to extract number from beginning or use default group
            try:
                # Try to extract number from start of string
                import re
                match = re.match(r'^(\d+)', str(schedule_id))
                if match:
                    schedule_num = int(match.group(1))
                    schedule_groups[schedule_num].append(row)
                else:
                    schedule_groups[0].append(row)
            except:
                schedule_groups[0].append(row)
    
    # Sort groups by schedule number
    sorted_groups = sorted(schedule_groups.items())
    
    # Debug output
    if not sorted_groups:
        print(f"  WARNING: No data grouped! Total records: {len(summary_data)}")
        if summary_data:
            print(f"  First record keys: {list(summary_data[0].keys())}")
            print(f"  First record values: {list(summary_data[0].values())}")
            # If no groups were created, create a single group with all data
            print(f"  Creating default group with all {len(summary_data)} records")
            sorted_groups = [(0, summary_data)]
    
    # Starting row for first schedule - row 6 (after header rows 2-4, with 1 empty row gap)
    current_row = 6
    
    # Debug: Print grouping info
    total_rows_to_write = sum(len(rows) for _, rows in sorted_groups)
    print(f"  Writing {total_rows_to_write} rows across {len(sorted_groups)} schedule group(s)")
    
    for schedule_num, rows in sorted_groups:
        # Sort rows by Schedule_ID or ID (case-insensitive)
        def get_sort_key(row):
            """Get sort key from row - tries Schedule_ID, then ID, then first column"""
            for key in ['Schedule_ID', 'SCHEDULE_ID', 'schedule_id', 'ScheduleID', 'ID', 'id']:
                if key in row:
                    return str(row[key])
            # Fallback to first column value
            return str(list(row.values())[0]) if row else ''
        rows.sort(key=get_sort_key)
        
        # Write Schedule title - left aligned, column A
        title_row = current_row
        title_text = schedule_titles.get(schedule_num, f"Schedule {schedule_num}")
        title_cell = ws.cell(row=title_row, column=1)
        title_cell.value = title_text
        title_cell.font = bold_font
        title_cell.alignment = Alignment(horizontal='left', vertical='top')
        
        # Write headers - Column A: ID, Column C: Description (B is gap), Column E: Value (D is gap)
        header_row = title_row + 1
        id_header = ws.cell(row=header_row, column=1)
        id_header.value = "ID"
        id_header.font = bold_font
        id_header.alignment = Alignment(horizontal='left', vertical='top')
        apply_border(ws, header_row, 1)
        
        desc_header = ws.cell(row=header_row, column=3)  # Column C (B is gap)
        desc_header.value = "Description"
        desc_header.font = bold_font
        desc_header.alignment = Alignment(horizontal='left', vertical='top')
        apply_border(ws, header_row, 3)
        
        # Column E: Value header removed - values will still be written in data rows
        # No header cell for Value column
        
        # Write data rows - 1 row after headers
        data_start_row = header_row + 1
        for idx, row_data in enumerate(rows):
            row_num = data_start_row + idx
            
            # Helper function for case-insensitive column access
            def get_col_value(row, col_name):
                """Get value from row with case-insensitive column name matching"""
                for k, v in row.items():
                    if k.upper() == col_name.upper():
                        return v
                return row.get(col_name, '')
            
            # Column A: Schedule_ID (e.g., "2-001") - also accepts ID, id
            schedule_id = get_col_value(row_data, 'Schedule_ID') or get_col_value(row_data, 'ID')
            id_cell = ws.cell(row=row_num, column=1)
            id_cell.value = schedule_id
            id_cell.font = default_font
            id_cell.alignment = Alignment(horizontal='left', vertical='top')
            apply_border(ws, row_num, 1)
            
            # Column C: Description (Column B is gap)
            description = get_col_value(row_data, 'Description')
            desc_cell = ws.cell(row=row_num, column=3)
            desc_cell.value = description
            desc_cell.font = default_font
            desc_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            apply_border(ws, row_num, 3)
            
            # Column E: Value (Column D is gap, highlighted in yellow)
            value = get_col_value(row_data, 'Value')
            value_cell = ws.cell(row=row_num, column=5)
            
            # Format numeric values with thousand separators (commas) - US format: 283,315
            # Try to convert to number if it's a string
            numeric_value = None
            if isinstance(value, (int, float)):
                numeric_value = value
            elif isinstance(value, str):
                # Try to convert string to number (remove commas if already formatted)
                try:
                    cleaned_value = value.replace(',', '').strip()
                    if cleaned_value:
                        numeric_value = float(cleaned_value) if '.' in cleaned_value else int(cleaned_value)
                except (ValueError, AttributeError):
                    pass
            
            if numeric_value is not None:
                # Set as number with Excel number format for comma separators
                value_cell.value = numeric_value
                value_cell.number_format = '#,##0'  # Excel format: 283,315
            else:
                # For non-numeric values like "N/A", keep as string
                value_cell.value = value
            
            value_cell.font = default_font
            value_cell.alignment = Alignment(horizontal='right', vertical='top')
            apply_border(ws, row_num, 5)
            apply_highlight(ws, row_num, 5)  # Yellow highlight
        
        # Move to next schedule section - leave 1 empty row between schedules
        current_row = data_start_row + len(rows) + 2  # +2 for empty row gap
    
    # Adjust column widths for better appearance
    ws.column_dimensions['A'].width = 15  # ID column
    ws.column_dimensions['B'].width = 3   # Gap column (narrow)
    ws.column_dimensions['C'].width = 70  # Description column
    ws.column_dimensions['D'].width = 3   # Gap column (narrow)
    ws.column_dimensions['E'].width = 18  # Value column
    
    # Set row heights
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    
    return ws


def create_worksheet(wb: Workbook, worksheet_config: WorksheetConfig,
                    detail_records: List[Dict[str, Any]], summaries: List[List[Dict[str, Any]]]):
    """Create a worksheet with detail and summary tables"""
    ws = wb.create_sheet(title=worksheet_config.name)
    
    # Note: Summary worksheet is handled separately in create_workbook
    # This function is only for detail worksheets
    if False:  # Placeholder - Summary is handled separately
        pass
    else:
        # Standard table layout
        # Determine header row
        header_row = worksheet_config.formatting.header_row if worksheet_config.formatting else 1
        
        # Check if this is a summary-only worksheet (state_summary_only template type)
        # For summary-only worksheets like 1-001 and 1-006, we skip writing detail records
        # The detail_records are still fetched and used to generate summaries, but not displayed
        # Detection: summary-only has summary_config, no spacing_columns, and detail_columns is None/empty
        is_summary_only = (worksheet_config.summary_config is not None and 
                          len(worksheet_config.spacing_columns) == 0 and
                          (worksheet_config.detail_columns is None or len(worksheet_config.detail_columns) == 0))
        
        # Write detail table only if not summary-only
        # Always write detail table (headers at minimum) even if no data
        if not is_summary_only:
            last_detail_row = write_detail_table(ws, detail_records, worksheet_config, start_row=header_row)
        else:
            # For summary-only worksheets, set last_detail_row to header_row
            # Summaries will be written starting from header_row
            last_detail_row = header_row
        
        # Handle spacing columns (spacing is handled by column positioning in summary_config)
        
        # Write summary tables (aligned with detail table header row, or start from header_row if no detail)
        if worksheet_config.summary_config and summaries:
            for summary_config, summary_data in zip(worksheet_config.summary_config, summaries):
                write_summary_table(ws, summary_data, summary_config, start_row=header_row)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def aggregate_all_worksheets_data(all_worksheets_data: Dict[str, Tuple[List[Dict[str, Any]], List[List[Dict[str, Any]]]]]) -> Dict[str, Any]:
    """Aggregate data from all detail worksheets for Summary sheet"""
    aggregated = {
        'all_records': [],
        'issue_state_counts': defaultdict(int),
        'resident_state_counts': defaultdict(int),
        'schedule_counts': defaultdict(int)
    }
    
    # Collect all records and aggregate counts
    for ws_name, (detail_records, summaries) in all_worksheets_data.items():
        if ws_name == "Summary":
            continue  # Skip Summary worksheet itself
        
        # Add all detail records
        aggregated['all_records'].extend(detail_records)
        
        # Aggregate Issue State and Resident State counts
        for record in detail_records:
            issue_state = record.get('Issue_State') or record.get('Issue_St') or record.get('issue_state') or ''
            resident_state = record.get('Resident_State') or record.get('Resident_St') or record.get('resident_state') or ''
            
            if issue_state:
                aggregated['issue_state_counts'][issue_state] += 1
            if resident_state:
                aggregated['resident_state_counts'][resident_state] += 1
        
        # Count records per schedule (worksheet name)
        aggregated['schedule_counts'][ws_name] = len(detail_records)
    
    return aggregated


def create_workbook(connection: snowflake.connector.SnowflakeConnection,
                   worksheets_config: List[WorksheetConfig],
                   summary_config: Dict[str, Any],
                   reporting_period: Optional[str] = None,
                   database: Optional[str] = None,
                   schema: Optional[str] = None) -> Workbook:
    """Create Excel workbook with all worksheets - Summary worksheet first"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Process Summary worksheet FIRST - query from summary table
    if summary_config and (summary_config.get('table_name') or summary_config.get('query')):
        summary_table_name = summary_config.get('table_name', '')
        schedule_titles = summary_config.get('schedule_titles', {})
        
        if summary_table_name:
            print(f"Processing Summary worksheet from table: {summary_table_name}")
        else:
            print(f"Processing Summary worksheet with custom query")
        
        # Check if custom query is provided
        custom_query = summary_config.get('query')
        filter_clause = summary_config.get('filter') or summary_config.get('where') or summary_config.get('where_clause')
        
        if custom_query:
            # Use custom query directly (supports multi-line SQL)
            query = custom_query.strip()
            print(f"  Using custom query for Summary worksheet")
        else:
            # Generate default query (requires table_name)
            if not summary_table_name:
                raise ValueError("Summary configuration requires either 'query' or 'table_name'")
            
            # Basic table name validation (alphanumeric, underscore, dot for schema.table)
            if not all(c.isalnum() or c in ('_', '.') for c in summary_table_name):
                raise ValueError(f"Invalid table name format: {summary_table_name}")
            
            # Build WHERE clause if filter is provided
            if filter_clause:
                where_clause = f"WHERE {filter_clause}"
            else:
                where_clause = ""
            
            query = f"SELECT Schedule_ID, Description, Value FROM {summary_table_name} {where_clause} ORDER BY Schedule_ID"
        
        summary_data = execute_query(connection, query, database, schema)
        print(f"  Fetched {len(summary_data)} summary records")
        
        # Debug: Print column names from first record if available
        if summary_data and len(summary_data) > 0:
            print(f"  Column names in query result: {list(summary_data[0].keys())}")
            # Check if required columns exist (case-insensitive)
            required_cols = ['Schedule_ID', 'Description', 'Value']
            actual_cols = [col.upper() for col in summary_data[0].keys()]
            missing_cols = [col for col in required_cols if col.upper() not in actual_cols]
            if missing_cols:
                print(f"  WARNING: Missing expected columns (case-insensitive): {missing_cols}")
                print(f"  Your query should return columns: Schedule_ID, Description, Value")
        else:
            print(f"  WARNING: No summary data returned from query!")
        
        # Convert schedule_titles from config (may be strings like "1", "2") to integers
        schedule_titles_dict = {}
        for key, value in schedule_titles.items():
            try:
                schedule_titles_dict[int(key)] = value
            except (ValueError, TypeError):
                schedule_titles_dict[key] = value
        
        create_summary_worksheet(wb, summary_data, schedule_titles_dict, reporting_period)
    
    # Process all detail worksheets AFTER Summary
    # Use parallel processing to speed up data fetching
    print(f"\nProcessing {len(worksheets_config)} worksheet(s)...")
    
    # Fetch data for all worksheets in parallel (I/O bound operation)
    def fetch_worksheet_data(ws_config):
        """Fetch data for a single worksheet"""
        try:
            detail_records, summaries = process_worksheet_data(connection, ws_config, database, schema)
            return ws_config.name, ws_config, detail_records, summaries
        except Exception as e:
            print(f"  Error processing {ws_config.name}: {e}")
            raise
    
    # Use ThreadPoolExecutor for parallel query execution
    # Note: Snowflake connection is thread-safe for read operations
    worksheet_data = {}  # Key: worksheet name (string), Value: (ws_config, detail_records, summaries)
    # Increase max_workers for better parallelization (up to 20 concurrent queries)
    # This helps when processing many worksheets
    max_workers = min(len(worksheets_config), 20)  # Increased from 10 to 20 for better performance
    
    print(f"  Using {max_workers} parallel workers for data fetching...")
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all worksheet data fetching tasks
        future_to_worksheet = {
            executor.submit(fetch_worksheet_data, ws_config): ws_config 
            for ws_config in worksheets_config
        }
        
        # Process completed tasks as they finish
        completed = 0
        for future in as_completed(future_to_worksheet):
            ws_config = future_to_worksheet[future]
            try:
                ws_name, ws_config, detail_records, summaries = future.result()
                worksheet_data[ws_name] = (ws_config, detail_records, summaries)
                completed += 1
                print(f"  [{completed}/{len(worksheets_config)}] Fetched data for {ws_name}: {len(detail_records)} records")
            except Exception as e:
                print(f"  Error fetching data for {ws_config.name}: {e}")
                raise
    
    # Write worksheets sequentially (Excel writing is not thread-safe)
    print(f"\nWriting worksheets to Excel...")
    for ws_config in worksheets_config:
        if ws_config.name in worksheet_data:
            ws_config, detail_records, summaries = worksheet_data[ws_config.name]
            print(f"  Writing worksheet: {ws_config.name}")
            print(f"    Detail records: {len(detail_records)}")
            print(f"    Summaries: {len(summaries) if summaries else 0}")
            if summaries:
                for idx, summary in enumerate(summaries):
                    print(f"      Summary {idx+1}: {len(summary)} rows")
            if ws_config.summary_config:
                print(f"    Summary configs: {len(ws_config.summary_config)}")
                for idx, sum_cfg in enumerate(ws_config.summary_config):
                    print(f"      Config {idx+1}: group_by='{sum_cfg.group_by}', start_column='{sum_cfg.start_column}'")
            create_worksheet(wb, ws_config, detail_records, summaries)
    
    return wb


# ============================================================================
# Main CLI Function
# ============================================================================

# ============================================================================
# Configuration Loading Functions
# ============================================================================

def load_config(config_path: str) -> Dict[str, Any]:
    """Load and parse YAML configuration file"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
        return config
    except FileNotFoundError:
        print(f"Error: Configuration file '{config_path}' not found.")
        sys.exit(1)
    except yaml.YAMLError as e:
        print(f"Error: Invalid YAML in configuration file: {e}")
        sys.exit(1)


def parse_config(config: Dict[str, Any]) -> Tuple[SnowflakeConfig, Dict[str, str], List[WorksheetConfig], Dict[str, Any]]:
    """
    Parse configuration: Snowflake settings, worksheet-to-table mapping, and summary config
    Snowflake connection parameters are read from environment variables first, then fall back to config.yaml
    """
    # Get Snowflake config from environment variables (preferred) or config file
    # Environment variables take precedence for security
    account = os.getenv('SNOWFLAKE_ACCOUNT') or (config.get('snowflake', {}).get('account') if 'snowflake' in config else None)
    user = os.getenv('SNOWFLAKE_USER') or (config.get('snowflake', {}).get('user') if 'snowflake' in config else None)
    password = os.getenv('SNOWFLAKE_PASSWORD') or (config.get('snowflake', {}).get('password') if 'snowflake' in config else None)
    warehouse = os.getenv('SNOWFLAKE_WAREHOUSE') or (config.get('snowflake', {}).get('warehouse') if 'snowflake' in config else None)
    database = os.getenv('SNOWFLAKE_DATABASE') or (config.get('snowflake', {}).get('database') if 'snowflake' in config else None)
    schema = os.getenv('SNOWFLAKE_SCHEMA') or (config.get('snowflake', {}).get('schema') if 'snowflake' in config else None)
    authenticator = os.getenv('SNOWFLAKE_AUTHENTICATOR') or (config.get('snowflake', {}).get('authenticator', 'externalbrowser') if 'snowflake' in config else 'externalbrowser')
    
    # Validate required fields
    required_fields = {
        'account': account,
        'warehouse': warehouse,
        'database': database,
        'schema': schema,
        'authenticator': authenticator
    }
    
    missing_fields = [field for field, value in required_fields.items() if not value]
    if missing_fields:
        raise ValueError(
            f"Missing required Snowflake configuration. "
            f"Set environment variables or config.yaml for: {', '.join(missing_fields)}. "
            f"Environment variables: SNOWFLAKE_ACCOUNT, SNOWFLAKE_WAREHOUSE, SNOWFLAKE_DATABASE, SNOWFLAKE_SCHEMA, SNOWFLAKE_AUTHENTICATOR"
        )
    
    # For username/password auth, validate credentials are provided
    if authenticator == 'snowflake':
        if not user or not password:
            raise ValueError(
                "For 'snowflake' authenticator, SNOWFLAKE_USER and SNOWFLAKE_PASSWORD "
                "must be set (via environment variables or config.yaml)"
            )
    
    snowflake_cfg = SnowflakeConfig(
        account=account,
        user=user,
        password=password,
        warehouse=warehouse,
        database=database,
        schema=schema,
        authenticator=authenticator
    )
    
    # Get worksheet configuration (can be string table_name or dict with table_name and template_type)
    worksheet_configs_raw = config.get('worksheets', {})
    worksheet_configs_raw = {k: v for k, v in worksheet_configs_raw.items() if k != 'Summary'}
    
    # Get summary configuration
    summary_config = config.get('summary', {})
    
    # Build worksheet configs
    worksheets_config = []
    for worksheet_name, worksheet_data in worksheet_configs_raw.items():
        # Support both formats:
        # 1. Simple: "1-001: table_name" (uses hardcoded structure)
        # 2. Advanced: "1-001: {table_name: 'table', template_type: 'direct_dump_state_summary'}" (uses template type)
        if isinstance(worksheet_data, str):
            # Simple format - use hardcoded structure
            table_name = worksheet_data
            ws_config = get_hardcoded_worksheet_structure(worksheet_name, table_name)
            if ws_config:
                worksheets_config.append(ws_config)
            else:
                print(f"Warning: Unknown worksheet name '{worksheet_name}', skipping...")
        elif isinstance(worksheet_data, dict) and 'table_name' not in worksheet_data and 'table' not in worksheet_data:
            # Simple format with just table name as dict value (backward compatibility)
            table_name = worksheet_data.get('table_name') or worksheet_data.get('table') or str(worksheet_data)
            ws_config = get_hardcoded_worksheet_structure(worksheet_name, table_name)
            if ws_config:
                worksheets_config.append(ws_config)
            else:
                print(f"Warning: Unknown worksheet name '{worksheet_name}', skipping...")
        elif isinstance(worksheet_data, dict):
            # Advanced format - use template type
            table_name = worksheet_data.get('table_name') or worksheet_data.get('table')
            template_type = worksheet_data.get('template_type')
            query = worksheet_data.get('query')
            detail_columns = worksheet_data.get('detail_columns')
            filter_clause = worksheet_data.get('filter') or worksheet_data.get('where') or worksheet_data.get('where_clause')
            currency_columns = worksheet_data.get('currency_columns')  # List of column names to format as currency
            
            if not table_name:
                print(f"Warning: Worksheet '{worksheet_name}' missing table_name, skipping...")
                continue
            
            if not template_type:
                print(f"Warning: Worksheet '{worksheet_name}' missing template_type, using hardcoded structure...")
                ws_config = get_hardcoded_worksheet_structure(worksheet_name, table_name)
                # Apply currency formatting to hardcoded config if specified
                if ws_config and currency_columns:
                    if ws_config.formatting:
                        ws_config.formatting.currency_columns = currency_columns
                    else:
                        ws_config.formatting = FormattingConfig(currency_columns=currency_columns)
            else:
                # Use template type to create config
                try:
                    ws_config = create_worksheet_config_from_template(
                        worksheet_name, table_name, template_type, query, detail_columns, filter_clause, currency_columns
                    )
                except ValueError as e:
                    print(f"Error: {e}")
                    print(f"  Falling back to hardcoded structure for '{worksheet_name}'...")
                    ws_config = get_hardcoded_worksheet_structure(worksheet_name, table_name)
            
            if ws_config:
                worksheets_config.append(ws_config)
        else:
            print(f"Warning: Invalid format for worksheet '{worksheet_name}', skipping...")
    
    # Build worksheet_tables mapping for backward compatibility (not used in current implementation)
    worksheet_tables = {}
    for ws_config in worksheets_config:
        # Extract table name from query if possible
        if ws_config.query:
            # Try to extract table name from query (simple pattern matching)
            match = re.search(r'FROM\s+([a-zA-Z_][a-zA-Z0-9_.]*)', ws_config.query, re.IGNORECASE)
            if match:
                worksheet_tables[ws_config.name] = match.group(1)
    
    return snowflake_cfg, worksheet_tables, worksheets_config, summary_config


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Generate multi-worksheet Excel report from Snowflake data'
    )
    parser.add_argument(
        '--config',
        type=str,
        required=True,
        help='Path to YAML configuration file with Snowflake settings and worksheet-to-table mapping'
    )
    parser.add_argument(
        '--output',
        type=str,
        required=True,
        help='Output Excel file path'
    )
    parser.add_argument(
        '--report-start-dt',
        type=str,
        required=True,
        help='Report start date (format: YYYY-MM-DD, MM/DD/YYYY, etc.)'
    )
    parser.add_argument(
        '--report-end-dt',
        type=str,
        required=True,
        help='Report end date (format: YYYY-MM-DD, MM/DD/YYYY, etc.)'
    )
    parser.add_argument(
        '--database',
        type=str,
        default=None,
        help='Override database for table name resolution in queries (optional)'
    )
    parser.add_argument(
        '--schema',
        type=str,
        default=None,
        help='Override schema for table name resolution in queries (optional)'
    )
    
    args = parser.parse_args()
    
    # Format dates and create reporting period string
    start_date_formatted = format_date_for_reporting_period(args.report_start_dt)
    end_date_formatted = format_date_for_reporting_period(args.report_end_dt)
    reporting_period = f"{start_date_formatted} through {end_date_formatted}"
    print(f"Reporting Period: {reporting_period}")
    
    # Load configuration
    print(f"Loading configuration from: {args.config}")
    config = load_config(args.config)
    
    # Parse configuration
    snowflake_cfg, worksheet_tables, worksheets_config, summary_config = parse_config(config)
    
    # Override database and schema for table name resolution if provided via command line
    # Note: Connection still uses config values, but queries will resolve table names using these
    target_database = args.database if args.database else snowflake_cfg.database
    target_schema = args.schema if args.schema else snowflake_cfg.schema
    
    if args.database or args.schema:
        print(f"Using command-line overrides for table name resolution:")
        if args.database:
            print(f"  Database: {target_database} (overridden from command line)")
        else:
            print(f"  Database: {target_database} (from config)")
        if args.schema:
            print(f"  Schema: {target_schema} (overridden from command line)")
        else:
            print(f"  Schema: {target_schema} (from config)")
    
    print(f"Found {len(worksheets_config)} detail worksheet(s) to process")
    if summary_config.get('table_name'):
        print(f"Summary worksheet will use table: {summary_config['table_name']}")
    
    # Create Snowflake connection (uses config values for connection)
    print("Connecting to Snowflake...")
    connection = create_snowflake_connection(snowflake_cfg)
    
    try:
        # Generate workbook (uses target_database/target_schema for table name resolution)
        print("Generating Excel workbook...")
        wb = create_workbook(connection, worksheets_config, summary_config, reporting_period,
                           database=target_database, schema=target_schema)
        
        # Save workbook
        print(f"Saving workbook to: {args.output}")
        wb.save(args.output)
        print(f"Successfully created Excel report: {args.output}")
        
    except Exception as e:
        print(f"Error generating workbook: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        close_connection(connection)


if __name__ == '__main__':
    main()

